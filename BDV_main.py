from data1 import QualityChecker
import csv 
import os
import cx_Oracle
from datetime import datetime
from sqlalchemy import create_engine
import pandas as pd
import numpy as np
from config_1 import Work_Folder_BDV,Reporting_folder_BDV,UPLOAD_FOLDER,BDV_Output
import openpyxl
from openpyxl import Workbook
import queries as qt

working_dir = Work_Folder_BDV
reporting_dir = Reporting_folder_BDV

file_path = Reporting_folder_BDV+'\\'+'FinalReport.xlsx'

if os.path.exists(file_path):
    os.remove(file_path)



# def bdv_validate(st,tt,conn_src,conn_tar):
#     error_dict={}
#     dfTargetTable=pd.DataFrame()

#     engine_src = create_engine(conn_src)
#     conn  = engine_src.connect()

#     engine_tar = create_engine(conn_tar)
#     conn_tar = engine_tar.connect()

#     try:

#         dfSourceTable=pd.read_sql(qt.GET_ALL_DATA_QUERY.format(st), conn)
#         dfTargetTable=pd.read_sql(qt.GET_ALL_DATA_QUERY.format(tt), conn_tar)
#         dfobj1 = pd.DataFrame()

#         src_len = len(dfSourceTable.index)

#         tar_len = len(dfTargetTable.index)

#         listIdenticalCol = [col for col in list(dfSourceTable.columns) if col in list(dfTargetTable.columns)]
#         dfDifference = pd.concat([dfSourceTable,dfTargetTable],sort=True)
#         dfDifference = dfDifference[listIdenticalCol]
#         dfDifference = dfDifference.reset_index(drop=True)
#         dfDifferencegrpby = dfDifference.groupby(list(dfDifference.columns))
#         idx = [x[0] for x in dfDifferencegrpby.groups.values() if len(x) == 1]
#         dfDifference=dfDifference.reindex(idx)

#         print("difference in the tables:",dfDifference)

#         if (src_len == tar_len):
#             return dfobj1,dfDifference,error_dict
#         else: 
#             return dfTargetTable,dfDifference,error_dict
#     except:
#         dfDifference=pd.DataFrame()
#         error_dict['error'] = "exception occured in unmatch_check"
#         return dfTargetTable,dfDifference,error_dict
            
# def datatype_check(st,tt,conn_src,conn_tar):
#     error_dict={}
#     df_datatype_diff = pd.DataFrame()

#     engine_src = create_engine(conn_src)
#     conn  = engine_src.connect()

#     engine_tar = create_engine(conn_tar)
#     conn_tar = engine_tar.connect()

#     try:   
#         dict_diff_datatype={}
#         dfSourceTable = pd.read_sql(qt.GET_ALL_DATA_QUERY.format(st),conn)
#         print("soruce table",dfSourceTable)
#         dfTargetTable = pd.read_sql(qt.GET_ALL_DATA_QUERY.format(tt),conn_tar)
#         listIdenticalCol = [col for col in list(dfSourceTable.columns) if col in list(dfTargetTable.columns)]
#         # print("listIdenticalCol", listIdenticalCol)
#         dfSourceCopy = dfSourceTable[listIdenticalCol].copy()
#         dfTargetCopy = dfTargetTable[listIdenticalCol].copy()
#         dict_src_col_type = dfSourceCopy.dtypes.astype(str).to_dict()
#         dict_target_col_type = dfTargetCopy.dtypes.astype(str).to_dict()
#         print("type targt dict",dict_target_col_type)
#         print("here")
        
#         for key,value in dict_target_col_type.items():
#             if dict_src_col_type[key] !=dict_target_col_type[key]:
#                 dict_diff_datatype[key] = value

#         if len(dict_diff_datatype)!=0:
#             print("hey")
#             df_datatype_diff = df_datatype_diff.append(dict_diff_datatype, ignore_index=True)
#             df_datatype_diff.insert(loc=0,column='Table_name',value=tt)

#         return st,df_datatype_diff,error_dict
        
#     except:        
#         error_dict["error"] = "exception occured in data type check"   
#         print (error_dict["error"])
#         return st,df_datatype_diff,error_dict



# def wrappr_BDV(src_engine=None,tar_engine=None,testcase_status_list=None,testcase_status_report=None):

# # path = "C:\\Users\\muhfnu\\Desktop\\Desktop_muhfnu\\DFTE_Demo\\Table_Details\\test_file.csv" 

#     path_src = UPLOAD_FOLDER + "\\source\\Tables_On_Premise_for_Metadata_Validation.csv"
#     path_tar = UPLOAD_FOLDER + "\\target\\Tables_On_Cloud_for_Metadata_Validation.csv"
#     response_dict = {}
#     response_dict["message"] = "status report"
#     response_dict["status"]="success"
    
#     testcase_status_list = []
#     testcase_status_report = []

#     k="Record Mismatch Validation"
#     k1="Record Count Validation"
#     k2="Table Structure Validation"

#     report_wb = Workbook()
#     report_sheet = report_wb.active
#     report_sheet.append(['Table','Records mismatch validation','Record Count validation','Table Structure Validation'])
#     report_wb.save(Reporting_folder_BDV+'\\'+'FinalReport.xlsx') 

#     st=[]
#     tt=[]


#     # table names from csv

#     with open(path_src) as csvfile:
#         readCSV = csv.reader(csvfile, delimiter=',')
#         for row in readCSV:
#                 st.append(row[0])


#     with open(path_tar) as csvfile:
#         readCSV = csv.reader(csvfile, delimiter=',')
#         for row in readCSV:
#                 tt.append(row[0])

#     for index, Source_tables in enumerate(st):
#         testcase_status = []
#         # testcase_status.append(st)

#         testcase_status_dict = {}
#         testcase_status_dict["Table"]=Source_tables
#         # print("testcase_status_dict",testcase_status_dict)

#         # report_incl = []

#         (df_a,df_b,error_dict) = bdv_validate(st[index],tt[index],src_engine,tar_engine)

#         df_len_row_cnt = len(df_a.index)
#         df_len_mismatch = len(df_b.index)

#         if df_len_row_cnt == 0:
#             if df_len_mismatch == 0:
#                 if len(error_dict) == 0:

#                     testcase_status.append("Pass")
#                     testcase_status_dict[str(k)]="Pass"
#                     testcase_status_dict[str(k1)]="Pass"
#                     testcase_status_dict[str(k2)]="Pass"
#                 else:
#                     testcase_status.append("Table not Available")
#                     testcase_status_dict[str(k)]="Table not Available"
#                     testcase_status_dict[str(k1)]="Table not Available"
#                     testcase_status_dict[str(k2)]="Table not Available"


#         else:
#             testcase_status.append("Fail")
#             testcase_status_dict[str(k)] = "Fail"
#             testcase_status_dict[str(k1)]="Fail"
#             testcase_status_dict[str(k2)]="Fail"

#         testcase_status_list.append(testcase_status_dict)
#         testcase_status_report.append(testcase_status)
#         create_folder(st[index])
#         write_data(st[index],datetime.now().strftime(st[index]+' %d-%m-%Y-%H-%M-%S.csv'),[df_a,df_b])

#         # report_incl_list.append(report_incl)
#     try:
#         response_dict = create_report(testcase_status_list,testcase_status_report)
#         # print("RESPONSE IN DO COMPARE",response_dict)
#         return response_dict
#     except Exception as e:
#         response_dict["error"] = str(e)
#         # response_dict["message"] = str(e)
#         return response_dict

# def create_folder(table_name):
#     if not(os.path.exists(os.path.join(BDV_Output,table_name))):
#         [os.makedirs(os.path.join(BDV_Output,table_name))]

# def write_data(table_name,file_name,df):
#     for ele in df:

#         my_list = ele.columns.values
#         my_list = [str(my_list[x]) for x in range(len(my_list))]
#         # print("my_list of column names",my_list)
#         header = ",".join(my_list)
#         # print("Header",header)

#         np.savetxt(os.path.join(BDV_Output,table_name,file_name),ele, delimiter=',',header=header,fmt='%s') 
#     # count=count+1

# def create_report(testcase_status_list=None,testcase_status_report=None):
#     response_dict = {}
#     response_dict["message"] = "status report"
#     response_dict["status"]="success"
#     response_dict["data"]= testcase_status_list
#     # print("RESPONSE DICT FOR REPORT",response_dict)
#     # print("TESTCASE_STATUS_REPORT",testcase_status_report)
#     try:
#         rep_wb = openpyxl.load_workbook(Reporting_folder_BDV + '\\' + 'FinalReport.xlsx')
#         rep_sheet = rep_wb.active
#         for tables in testcase_status_list:
#             rep_sheet.append(tables)
#         rep_wb.save(Reporting_folder_BDV + '\\' + 'FinalReport.xlsx')
#     except:
#         response_dict["message"]="Exception occured, could not save the test report."
#     # except Exception as e:
#     #     response_dict["error"] = str(e)
#     #     # response_dict["message"] = str(e)
#     #     return response_dict

#     return response_dict

# //////////////////////////////////////////////////////////////////////////////////////////////////////



def do_compare1(data, staging_dir):
    response_dict={}
    # pass the connection details from quality checker
    testcase_status_list = []
    testcase_status_report = []
    quality_checker = QualityChecker(data) 
    st = []
    tt = []

    # test cases  

    function_mapping = {
        "Table Structure Validation" : quality_checker.datatype_check,
        "Record Count validation" : quality_checker.row_count_check,
        "Records mismatch validation" : quality_checker.unmatch_check,
        }
    
    #  Creating a Subset of function mapping using the UI request

    # REQ_OBJECT = {"Validations":["Table Structure Validation","Sequences Validation","Constraints Validation","Triggers Validation"]}
    REQ_OBJECT = data["Validations"]
    print("REQ_OBJECT",REQ_OBJECT)

    function_mapping_subset={}
    for values in REQ_OBJECT:
        # print("v",values)
        for k,v in function_mapping.items():
                if values==k:
                        function_mapping_subset.update({k:v})
        
    # print("subset",function_mapping_subset)
   
    # read source CSV files
    try:
        with open(os.path.join(staging_dir, "source", data["source_file_name"])) as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                st.append(row[0])
    except Exception as e:
        response_dict["error"] = str(e)
        response_dict["message"] = str(e)
        return response_dict
        
    # read target CSV files    
    #     
    try:
        with open(os.path.join(staging_dir, "target", data["target_file_name"])) as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                tt.append(row[0])
    except Exception as e:
        response_dict["error"] = str(e)
        response_dict["message"] = str(e)
        return response_dict
    
#  making the final report

    report_wb = Workbook()
    report_sheet = report_wb.active
    report_sheet.append(['Tables', 'Records mismatch validation','Record Count validation','Table Structure Validation'])
    report_wb.save(reporting_dir+'\\'+'FinalReport.xlsx') 

# Running the test validation

    for index, source_table in enumerate(st):
        
        
        try:
            
            compare_tables(source_table,tt[index],function_mapping,testcase_status_list,testcase_status_report)
            # except Exception as e:
        # except:
        #     pass
        # #         # response_dict["error"] = str(e)
        # #         # response_dict["message"] = str(e)
        # #         # return response_dict

        except Exception as e:
            print (e)
            response_dict["error"] = str(e)
            return response_dict
        
        # finally:
        #     pass



#  Response for the status table in the UI

    try:
        response_dict = create_report(testcase_status_list,testcase_status_report)
        # print("RESPONSE IN DO COMPARE",response_dict)
        return response_dict
    except Exception as e:
        response_dict["error"] = str(e)
        # response_dict["message"] = str(e)
        return response_dict
    


    
def compare_tables(st, tt, function_mapping,testcase_status_list=None,testcase_status_report=None):
    
    testcase_status = []
    testcase_status.append(st)
    testcase_status_dict = {}
    testcase_status_dict["Table"]=st
    # testcase_status.append(st)
    for k, v in function_mapping.items():
        print("Comparing tables : ",st +" from On-Premise Database ",tt +" from Cloud Database ")
        print("\n")
        print("Performing :", k)
        print("\n")
        (table_name, df,error_dict) = v(st, tt)
        print("\n")
        # print("The Case is Tested and results are stored in :",working_dir)
        create_folder(table_name, str(k))
        # print("dataframe",df)
        if df.empty:
            if len(error_dict) != 0:
                # print("error dict",error_dict)
                testcase_status.append("Table not available in Target")
                testcase_status_dict[str(k)] = "Table not available in Target"
            else:
                testcase_status.append("Pass")
                testcase_status_dict[str(k)]="Pass"
        else:
            if len(error_dict) != 0:
                testcase_status_dict[str(k)] = "error"
                testcase_status.append("error")
            else:
                # testcase_status.append("Fail")
                testcase_status.append("Fail")
                testcase_status_dict[str(k)] = "Fail"

            write_data(table_name, str(k), datetime.now().strftime(table_name+' %d-%m-%Y-%H-%M-%S.csv'), df)
    testcase_status_list.append(testcase_status_dict)
    testcase_status_report.append(testcase_status)
    # creating report in the local system
    
    
  
        # creating report in the local system
def create_report(testcase_status_list=None,testcase_status_report=None):
    response_dict = {}
    response_dict["message"] = "status report"
    response_dict["status"]="success"
    response_dict["data"]= testcase_status_list
    # print("RESPONSE DICT FOR REPORT \n",response_dict)
    # print("TESTCASE_STATUS_REPORT \n",testcase_status_report)
    print("The status report is saved in the static Directory")
    try:
        rep_wb = openpyxl.load_workbook(reporting_dir + '\\' + 'FinalReport.xlsx')
        rep_sheet = rep_wb.active
        for tables in testcase_status_report:
            rep_sheet.append(tables)
        rep_wb.save(reporting_dir + '\\' + 'FinalReport.xlsx')
    except:
        response_dict["message"]="Exception occured, could not save the test report."
    return response_dict
  
  
  
  
# def create_report_1(testcase_status=None):
    
#     rep_wb = openpyxl.load_workbook(reporting_dir+'\\'+'FinalReport.xlsx')
#     rep_sheet = rep_wb.active
#     rep_sheet.append(testcase_status)
#     rep_wb.save(reporting_dir+'\\'+'FinalReport.xlsx')

# #  report view in the UI
# def create_report(testcase_status_list=None):
#     response_dict = {}
#     response_dict["message"] = "status report"
#     response_dict["status"]="success"
#     response_dict["data"]= testcase_status_list
#     print("RESPONSE DICT FOR REPORT",response_dict)
#     return response_dict



def create_folder(table_name, check_name):
    if not(os.path.exists(os.path.join(working_dir, table_name, check_name))):
        os.makedirs(os.path.join(working_dir, table_name, check_name))
    
def write_data(folder_name, check_name, file_name, df):
    my_list = df.columns.values
    my_list = [str(my_list[x]) for x in range(len(my_list))]
    # print (my_list)
    header = ",".join(my_list)
    np.savetxt(os.path.join(working_dir, folder_name, check_name,  file_name), df , header = header,   delimiter=',', fmt='%s')
    



