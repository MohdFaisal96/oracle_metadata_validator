from data1 import QualityChecker
import os
import numpy as np
import csv
from os import listdir
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from config_1 import UPLOAD_FOLDER,Work_Folder,Reporting_folder
from data1 import logger

st = []
tt = []


working_dir = Work_Folder
reporting_dir = Reporting_folder
file_path = reporting_dir + '\\' + 'FinalReport.xlsx'

# reporting_dir = "C:\\Users\\muhfnu\\Desktop\\Desktop_muhfnu\\DFTE_Demo\\Table_Details"
# file_path = "C:\\Users\\muhfnu\\Desktop\\Desktop_muhfnu\\DFTE_Demo\\Table_Details\\FinalReport.xlsx"
# working_dir = "C:\\Users\\muhfnu\\Desktop\\Desktop_muhfnu\\DFTE_Demo\\Table_Details\\output"

if os.path.exists(file_path):
    os.remove(file_path)


# def do_compare(data, staging_dir):

#     # pass the connection details to quality checker
    
#     quality_checker = QualityChecker(data)

    
#     st = []
#     tt = []

#     function_mapping = {
    
#         "Records mismatch validation" : quality_checker.unmatch_check,
#         "Constraints Validation" : quality_checker.constraint_check,
#         "Record Count validation" : quality_checker.row_count_check,
#         "Table Structure Validation" : quality_checker.datatype_check
        
#         }
    
   
#     # read CSV files
#     with open(os.path.join(staging_dir, "source", data["source_file_name"])) as csvfile:
        
#         readCSV = csv.reader(csvfile, delimiter=',')
#         for row in readCSV:
#             st.append(row[0])
            
            
#     with open(os.path.join(staging_dir, "target", data["target_file_name"])) as csvfile:
#         readCSV = csv.reader(csvfile, delimiter=',')
#         for row in readCSV:
#             tt.append(row[0])
     
#             #  Status Report
            
#     report_wb = Workbook()
#     report_sheet = report_wb.active
#     report_sheet.append(['Tables', 'Records mismatch validation','Constraints Validation','Record Count validation','Table Structure Validation'])
#     report_wb.save(reporting_dir+'\\'+'FinalReport.xlsx')
            
#     for index, source_table in enumerate(st):
#         compare_tables(source_table,tt[index], function_mapping)
    
    
#     # return data
    
    
    
    
# def compare_tables(st, tt, function_mapping):
#     testcase_status = []
#     testcase_status.append(st)
#     for k, v in function_mapping.items():
#         print("Comparing tables : ",st +" from On-Premise Database ",tt +" from Cloud Database ")
#         print("\n")
#         print("Performing :", k)
#         print("\n")
#         (table_name, df) = v(st, tt)
#         print("\n")
#         print("The Case is Tested and results are stored in :",working_dir)
#         create_folder(table_name , str(k))
            
            
#         if df.empty:
#             testcase_status.append("Pass")
#         else:
#             testcase_status.append("Fail")
#             print("dict_testcase",testcase_status)
#             write_data(table_name, str(k), datetime.now().strftime(table_name+' %d-%m-%Y-%H-%M-%S.csv'), df)
#     create_report(testcase_status)
    
  
        
# def create_report(testcase_status=None):
    
#     rep_wb = openpyxl.load_workbook(reporting_dir+'\\'+'FinalReport.xlsx')
#     rep_sheet = rep_wb.active
#     rep_sheet.append(testcase_status)
#     rep_wb.save(reporting_dir+'\\'+'FinalReport.xlsx')
          
# def create_folder(table_name, check_name):
#     if not(os.path.exists(os.path.join(working_dir, table_name, check_name))):
#         os.makedirs(os.path.join(working_dir, table_name, check_name))
    
# def write_data(folder_name, check_name, file_name, df):
#     my_list = df.columns.values
#     my_list = [str(my_list[x]) for x in range(len(my_list))]
#     # print (my_list)
#     header = ",".join(my_list)
#     np.savetxt(os.path.join(working_dir, folder_name, check_name,  file_name), df , header = header,   delimiter=',', fmt='%s')

# //////////// NOT REqUIRED\\\\\\\\\\\\\\\\\\\\\\
# def main():
#     for index, source_table in enumerate(st):
#         compare_tables(source_table,tt[index])

# if __name__ == "__main__":
#     main()
# close_var = raw_input("Exit")
 
#  ////////////////////////////////////// NEW UPDATE?\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\
#     \\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\///////////////////////////////////////////////////////
    
def do_compare(data, staging_dir):
    response_dict={}
    # pass the connection details from quality checker
    testcase_status_list = []
    testcase_status_report = []
    quality_checker = QualityChecker(data)

    
    st = []
    tt = []

    # test cases

    
    

    function_mapping = {
        "Table Structure Validation" : quality_checker.datatype_check,
        "Constraints Validation" : quality_checker.constraint_check,
        "Triggers Validation" : quality_checker.trigger_check,
        "Sequences Validation" : quality_checker.Sequence_check,
        "Record Count validation" : quality_checker.row_count_check,
        "Records mismatch validation" : quality_checker.unmatch_check,
        "Index Validation" : quality_checker.index_check
        }
    
    #  Creating a Subset of function mapping using the UI request

    # REQ_OBJECT = {"Validations":["Table Structure Validation","Sequences Validation","Constraints Validation","Triggers Validation"]}
    REQ_OBJECT = data["Validations"]
    logger.info("Im in DataValidtion")
    logger.info(REQ_OBJECT)

    # user_function=[]
    # for k,v in REQ_OBJECT.items():
    #     for v in REQ_OBJECT["Validations"]:
    #         # print("Key",k + " Value",v)
    #         user_function.append(v)
      
    # # print("subset",user_function)
    

    # for i in user_function:
    #     for k,v in function_mapping.items():
    #         if i==k:
    #             # print("v",v)
    #             function_mapping_subset.update({k:v})

    function_mapping_subset={}
    for values in REQ_OBJECT:
        # print("v",values)
        for k,v in function_mapping.items():
                if values==k:
                        function_mapping_subset.update({k:v})
        
    print("subset >>>>>>>>>  ",function_mapping_subset)
   
    # read source CSV files
    try:
        with open(os.path.join(staging_dir, "source", data["source_file_name"])) as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                st.append(row[0])
    except Exception as e:
        response_dict["error"] = str(e)
        response_dict["message"] = str(e)
        return response_dict
        
    # read target CSV files    
    #     
    try:
        with open(os.path.join(staging_dir, "target", data["target_file_name"])) as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                tt.append(row[0])
    except Exception as e:
        response_dict["error"] = str(e)
        response_dict["message"] = str(e)
        return response_dict
    
#  making the final report

    header_list = []

    # for keys in function_mapping_subset:
    #     header_list.append(keys)

    # print("header list >>>",header_list)


    report_wb = Workbook()
    report_sheet = report_wb.active

    # report_sheet.append(header_list)
    
    # report_sheet.append(['Tables', 'Records mismatch validation','Constraints Validation','Record Count validation','Table Structure Validation','Trigger Validation'])
   
    for keys in function_mapping_subset:

        report_sheet.append(['Tables',keys])
    report_wb.save(reporting_dir+'\\'+'FinalReport.xlsx') 

# Running the test validation

    for index, source_table in enumerate(st):
        
        
        try:
            
            compare_tables(source_table,tt[index],function_mapping_subset,testcase_status_list,testcase_status_report)
            # except Exception as e:
        # except:
        #     pass
        # #         # response_dict["error"] = str(e)
        # #         # response_dict["message"] = str(e)
        # #         # return response_dict

        except Exception as e:
            # print (e)
            response_dict["error"] = str(e)
            return response_dict
        
        # finally:
        #     pass



#  Response for the status table in the UI

    try:
        response_dict = create_report(testcase_status_list,testcase_status_report)
        # print("RESPONSE IN DO COMPARE",response_dict)
        return response_dict
    except Exception as e:
        response_dict["error"] = str(e)
        # response_dict["message"] = str(e)
        return response_dict
    


    
def compare_tables(st, tt, function_mapping,testcase_status_list=None,testcase_status_report=None):
    
    testcase_status = []
    testcase_status.append(st)
    testcase_status_dict = {}
    testcase_status_dict["Table"]=st
    # testcase_status.append(st)
    for k, v in function_mapping.items():
        print("Comparing tables : ",st +" from On-Premise Database ",tt +" from Cloud Database ")
        print("\n")
        print("Performing :", k)
        print("\n")
        (table_name, df,error_dict) = v(st, tt)
        print("\n")
        # print("The Case is Tested and results are stored in :",working_dir)
        create_folder(table_name, str(k))
        # print("dataframe",df)
        if df.empty:
            if len(error_dict) != 0:
                # print("error dict",error_dict)
                testcase_status.append("Table not available in Target")
                testcase_status_dict[str(k)] = "Table not available in Target"
            else:
                testcase_status.append("Pass")
                testcase_status_dict[str(k)]="Pass"
        else:
            if len(error_dict) != 0:
                testcase_status_dict[str(k)] = "error"
                testcase_status.append("error")
            else:
                # testcase_status.append("Fail")
                testcase_status.append("Fail")
                testcase_status_dict[str(k)] = "Fail"

            write_data(table_name, str(k), datetime.now().strftime(table_name+' %d-%m-%Y-%H-%M-%S.csv'), df)
    testcase_status_list.append(testcase_status_dict)
    testcase_status_report.append(testcase_status)
    # creating report in the local system
    
    
  
        # creating report in the local system
def create_report(testcase_status_list=None,testcase_status_report=None):
    response_dict = {}
    response_dict["message"] = "status report"
    response_dict["status"]="success"
    response_dict["data"]= testcase_status_list
    # print("RESPONSE DICT FOR REPORT \n",response_dict)
    # print("TESTCASE_STATUS_REPORT \n",testcase_status_report)
    print("The status report is saved in the static Directory")
    try:
        rep_wb = openpyxl.load_workbook(reporting_dir + '\\' + 'FinalReport.xlsx')
        rep_sheet = rep_wb.active
        for tables in testcase_status_report:
            rep_sheet.append(tables)
        rep_wb.save(reporting_dir + '\\' + 'FinalReport.xlsx')
    except:
        response_dict["message"]="Exception occured, could not save the test report."
    return response_dict
  
  
  
  
# def create_report_1(testcase_status=None):
    
#     rep_wb = openpyxl.load_workbook(reporting_dir+'\\'+'FinalReport.xlsx')
#     rep_sheet = rep_wb.active
#     rep_sheet.append(testcase_status)
#     rep_wb.save(reporting_dir+'\\'+'FinalReport.xlsx')

# #  report view in the UI
# def create_report(testcase_status_list=None):
#     response_dict = {}
#     response_dict["message"] = "status report"
#     response_dict["status"]="success"
#     response_dict["data"]= testcase_status_list
#     print("RESPONSE DICT FOR REPORT",response_dict)
#     return response_dict



def create_folder(table_name, check_name):
    if not(os.path.exists(os.path.join(working_dir, table_name, check_name))):
        os.makedirs(os.path.join(working_dir, table_name, check_name))
    
def write_data(folder_name, check_name, file_name, df):
    my_list = df.columns.values
    my_list = [str(my_list[x]) for x in range(len(my_list))]
    # print (my_list)
    header = ",".join(my_list)
    np.savetxt(os.path.join(working_dir, folder_name, check_name,  file_name), df , header = header,   delimiter=',', fmt='%s')
    
